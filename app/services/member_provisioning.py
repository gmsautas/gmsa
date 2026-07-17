"""Shared member-provisioning helpers: find-or-create a User (normally set up
via an emailed one-click link; optionally via a real temp password + the same
link handed back for offline distribution instead) and generic CSV/XLSX row
parsing for bulk imports.

Used by both the election voter-register upload (app.services.elections) and
the standalone member bulk-upload (app.web.admin_web) — kept in its own
module, separate from app.services.elections, so provisioning a plain member
account doesn't require importing the election machinery.
"""

import csv
import io
import re
import secrets
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.models import User
from app.services import academic, email_failures, password_reset, resend_client


def extract_first_name(name: str) -> str:
    """Best-effort "first name" for a spreadsheet column, out of the single
    free-text `name` field this app stores (there's no separate first/last
    name on User) -- just the first whitespace-separated token."""
    parts = (name or "").split()
    return parts[0] if parts else (name or "")


class ProvisioningConflict(Exception):
    """Raised when a row's email and student_id resolve to two different existing accounts."""


def normalize_phone(raw: str) -> str:
    """Light cleanup for real-world register phone columns: drops values
    that are actually an email pasted into the phone column (a copy-paste
    mistake we've seen in real registers), and fixes the common "O" (letter)
    for "0" (digit) typo at the start of Ghanaian numbers. Doesn't otherwise
    reformat/validate -- e.g. two numbers separated by "/" are left as-is
    rather than guessing which one to drop."""
    value = (raw or "").strip()
    if not value or "@" in value:
        return ""
    rest = value[1:].replace(" ", "").replace("/", "")
    if value[:1] in ("O", "o") and rest.isdigit():
        value = "0" + value[1:]
    return value


async def _find_existing_member(db: AsyncSession, *, student_id: str, email: str) -> User | None:
    """Shared lookup/conflict-check core of find_or_create_member and
    find_or_create_member_no_email -- looks up by email and by student_id and
    raises ProvisioningConflict if they resolve to two different accounts (or
    to an existing account whose student_id doesn't match this row's).

    Raises ProvisioningConflict if the email and student_id belong to two
    different existing accounts -- this also catches two different students
    sharing one email within the same import batch (the earlier row's
    account is flushed and visible to this query before the batch's final
    commit), which otherwise would silently merge the second student into
    the first student's account instead of giving them their own.
    """
    user_by_email = (
        await db.execute(select(User).where(func.lower(User.email) == email))
    ).scalar_one_or_none()
    user_by_sid = (
        await db.execute(select(User).where(User.student_id == student_id))
    ).scalar_one_or_none()

    if user_by_email and user_by_sid and user_by_email.id != user_by_sid.id:
        raise ProvisioningConflict(
            f"Email matches account #{user_by_email.id} but student ID matches "
            f"a different account #{user_by_sid.id}"
        )

    if user_by_email is not None and user_by_email.student_id and user_by_email.student_id != student_id:
        raise ProvisioningConflict(
            f"Email {email} already belongs to account #{user_by_email.id} "
            f"(student ID {user_by_email.student_id}), but this row's student ID is "
            f"{student_id} -- likely two different students sharing one email by "
            f"mistake in the source file; fix one of them and re-import that row"
        )

    return user_by_email or user_by_sid


async def find_or_create_member(
    db: AsyncSession, *, student_id: str, email: str, name: str, base_url: str, send_setup_email: bool = True
) -> tuple[User, bool, bool, str | None]:
    """Find an existing user by email or student_id, or create one and issue
    it a one-click account-setup link.

    Returns (user, created, email_failed, account_setup_url).

    When send_setup_email=True (default, used by import_members): the setup
    link is emailed immediately for a newly created user; email_failed
    reflects whether that send succeeded, and account_setup_url is always
    None.

    When send_setup_email=False (used by elections.import_register): the
    link is still issued and durably committed, but NOT emailed --
    email_failed is always False here, and account_setup_url holds the link
    so the caller can send it itself, e.g. combined with a voter token into
    one email instead of two separate ones. The caller becomes responsible
    for handling/recording a failed send of that link.

    Raises ProvisioningConflict -- see _find_existing_member.
    """
    user = await _find_existing_member(db, student_id=student_id, email=email)
    if user is not None:
        return user, False, False, None

    user = User(
        name=name or email.split("@")[0],
        email=email,
        student_id=student_id,
        # Placeholder only -- nobody is ever told this value. The account is
        # activated by setting a real password through the emailed link.
        password_hash=await hash_password(secrets.token_urlsafe(32)),
        must_change_password=True,
        role="member",
        status="active",
    )
    db.add(user)
    await db.flush()

    email_failed = False
    account_setup_url = None
    if send_setup_email:
        # send_account_setup_link -> _issue_and_email_link commits (the new
        # token row, plus this just-flushed user in the same transaction)
        # before it sends the email, so the user is already durable by the
        # time the email goes out.
        try:
            await password_reset.send_account_setup_link(db, user=user, base_url=base_url)
        except resend_client.ResendError as err:
            email_failed = True
            await email_failures.record_failure(
                db, recipient=user.email, purpose="account_setup", error=err
            )
    else:
        # Same durability guarantee (issue_account_setup_link commits before
        # returning) -- just doesn't send anything. The caller sends it,
        # combined or not, and is responsible for handling that send's
        # failure.
        account_setup_url = await password_reset.issue_account_setup_link(
            db, user=user, base_url=base_url
        )

    return user, True, email_failed, account_setup_url


async def find_or_create_member_no_email(
    db: AsyncSession, *, student_id: str, email: str, name: str, base_url: str
) -> tuple[User, bool, str | None, str | None]:
    """Find-or-create variant for imports run with sending disabled entirely
    (see elections.import_register's and import_members's send_emails=False
    path) -- no email goes out. A newly created account instead gets BOTH:

    - a real, admin-visible temporary password (same style as
      reset_password_for_admin_reveal) for handing to the member verbally
      (phone, in person, WhatsApp text), and
    - a one-click account-setup link (same mechanism as
      find_or_create_member's emailed link, just not emailed here) for
      handing to them as something to click instead -- following it lands on
      the same "choose your password" page, and now auto-logs them straight
      in on success (see app.web.auth_web.reset_password_submit).

    Both are meant to be bundled into a credentials spreadsheet (see
    build_credentials_workbook) for the admin to distribute offline.

    Returns (user, created, temp_password, setup_url) -- temp_password and
    setup_url are both None when the account already existed, since
    importing a register never resets an existing member's credentials.

    Raises ProvisioningConflict -- see _find_existing_member.
    """
    user = await _find_existing_member(db, student_id=student_id, email=email)
    if user is not None:
        return user, False, None, None

    temp_password = secrets.token_urlsafe(9)
    user = User(
        name=name or email.split("@")[0],
        email=email,
        student_id=student_id,
        password_hash=await hash_password(temp_password),
        must_change_password=True,
        role="member",
        status="active",
    )
    db.add(user)
    await db.flush()

    # Commits (the new token row, plus this just-flushed user) before
    # returning -- same durability guarantee as find_or_create_member's
    # issue_account_setup_link call, just never emailed.
    setup_url = await password_reset.issue_account_setup_link(db, user=user, base_url=base_url)

    return user, True, temp_password, setup_url


async def reset_and_resend_welcome_email(db: AsyncSession, user: User, *, base_url: str) -> None:
    """For an existing user whose original account-setup email never
    arrived. Issues a fresh setup link (invalidating any earlier one) and
    emails it. Raises ResendError on a failed send so callers (e.g.
    scripts/local_gmail_import.py's --resend-account-email-for) can detect
    and report it rather than wrongly assuming success."""
    await password_reset.send_account_setup_link(db, user=user, base_url=base_url)


async def reset_password_for_admin_reveal(db: AsyncSession, user: User) -> tuple[str, bool]:
    """Issue a fresh temp password for an admin to read off screen and hand
    to the member directly (phone call, in person, WhatsApp) -- for when
    email delivery can't be relied on at all, so there's no dependency on it
    working. Deliberately password-based (not a link) since that's what's
    actually usable when reading off screen or over the phone. Always
    persists the new password and never raises on a failed send (that's the
    whole point); returns (temp_password, email_sent) so the caller can still
    show whether the notification email also went out."""
    temp_password = secrets.token_urlsafe(9)
    user.password_hash = await hash_password(temp_password)
    user.must_change_password = True
    await db.commit()

    email_sent = True
    try:
        html = (
            f"<p>Assalamu alaikum {user.name},</p>"
            f"<p>Your GMSA UTAS account password was reset by an administrator.</p>"
            f"<p>Email: <strong>{user.email}</strong><br>"
            f"Temporary password: <strong>{temp_password}</strong></p>"
            f"<p>Please log in and you will be asked to set a new password before doing anything "
            f"else.</p>"
        )
        await resend_client.send_email(
            to=[user.email], subject="Your GMSA UTAS password was reset", html=html
        )
    except resend_client.ResendError as err:
        email_sent = False
        await email_failures.record_failure(
            db, recipient=user.email, purpose="admin_password_reveal_notification", error=err
        )

    return temp_password, email_sent


@dataclass
class MemberImportResult:
    created_users: int = 0
    linked_users: int = 0
    skipped_duplicates: int = 0
    conflicts: list[dict] = field(default_factory=list)
    email_failures: list[str] = field(default_factory=list)
    # Populated only when import_members is run with send_emails=False -- one
    # entry per newly created account, holding everything an admin needs to
    # hand a member their credentials offline instead of by email. See
    # build_credentials_workbook.
    credentials: list[dict] = field(default_factory=list)


async def import_members(
    db: AsyncSession,
    rows: list[dict],
    *,
    base_url: str,
    send_emails: bool = True,
    on_row: Callable[[int, int, dict], None] | None = None,
) -> MemberImportResult:
    """Bulk create/link member accounts from parsed CSV/XLSX rows — the
    standalone counterpart to app.services.elections.import_register, minus
    any Voter/election coupling. Recognizes optional phone/program/
    program_category/level columns, applied only to newly created accounts
    (an existing account's details are never overwritten by a bulk upload).

    send_emails=False turns off the account-setup email entirely and issues
    each newly created account a real temp password + setup link instead
    (see find_or_create_member_no_email), collected into result.credentials
    for the caller to offer as a downloadable spreadsheet. Existing/linked
    accounts are never touched either way, so they never get a credentials
    row.

    `on_row`, if given, is called synchronously after each row finishes
    processing as `on_row(index, total, info)` (1-based index) -- same
    contract as elections.import_register's on_row, minus the voter-specific
    keys (there's no Voter/election here). `info` always has "email" and
    "outcome" ("imported" | "skipped_duplicate" | "conflict") plus, when
    "outcome" is "imported": "created" (bool) and "email_failed" (bool,
    always False when created is False -- an existing/linked account is never
    emailed by this import).
    """
    result = MemberImportResult()
    seen: set[tuple[str, str]] = set()
    total = len(rows)

    for index, row in enumerate(rows, start=1):
        student_id = str(row.get("student_id") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        name = str(row.get("name") or "").strip()
        phone = normalize_phone(str(row.get("phone") or ""))
        program = str(row.get("program") or "").strip()
        program_category = str(row.get("program_category") or "").strip().lower()
        level_raw = str(row.get("level") or "").strip()

        if not student_id or not email:
            result.conflicts.append({"row": row, "reason": "Missing student_id or email"})
            if on_row:
                on_row(index, total, {"email": email, "outcome": "conflict"})
            continue

        key = (student_id, email)
        if key in seen:
            result.skipped_duplicates += 1
            if on_row:
                on_row(index, total, {"email": email, "outcome": "skipped_duplicate"})
            continue
        seen.add(key)

        temp_password = None
        setup_url = None
        try:
            if send_emails:
                user, created, email_failed, _account_setup_url = await find_or_create_member(
                    db, student_id=student_id, email=email, name=name, base_url=base_url
                )
            else:
                email_failed = False
                user, created, temp_password, setup_url = await find_or_create_member_no_email(
                    db, student_id=student_id, email=email, name=name, base_url=base_url
                )
        except ProvisioningConflict as err:
            result.conflicts.append({"row": row, "reason": str(err)})
            if on_row:
                on_row(index, total, {"email": email, "outcome": "conflict"})
            continue

        if created:
            result.created_users += 1
            if phone:
                user.phone = phone
            if program:
                user.program = program
            if program_category in academic.PROGRAM_CATEGORIES:
                user.program_category = program_category
                user.grad_year = academic.graduation_year(student_id, program_category)
            if level_raw:
                try:
                    user.level_override = int(level_raw)
                except ValueError:
                    pass
            if email_failed:
                result.email_failures.append(f"account email to {user.email}")
            if not send_emails:
                result.credentials.append(
                    {
                        "first_name": extract_first_name(user.name),
                        "email": user.email,
                        "phone": user.phone or "",
                        "temp_password": temp_password or "",
                        "setup_link": setup_url or "",
                    }
                )
        else:
            result.linked_users += 1

        # Commit per row rather than once at the end -- a large member
        # import runs as one long web request, so this keeps a request
        # that gets cut off partway from rolling back rows already done.
        await db.commit()

        if on_row:
            on_row(
                index,
                total,
                {
                    "email": email,
                    "outcome": "imported",
                    "created": created,
                    "email_failed": email_failed if created else False,
                },
            )

    return result


def build_credentials_workbook(credentials: list[dict], *, include_voter_token: bool = False) -> bytes:
    """Renders a send_emails=False result's credentials (import_members's or
    app.services.elections.import_register's) into an .xlsx file for the
    admin to download once, in place of the account-setup/voter-token emails
    that were skipped. Never written to disk by the caller -- generated on
    demand from data already held in memory for this one request/response.

    include_voter_token=True adds the extra "Voter Token" column the election
    register import needs (each credentials dict must then also have a
    "voter_token" key); plain member imports have no election/voter
    involved, so leave it False.

    The "Account Setup Link" column, when present, is rendered as a real
    clickable hyperlink (not just plaintext) -- following it lands the
    member/voter on the same "choose your password" page the emailed link
    would have, and now auto-logs them in on success (see
    app.web.auth_web.reset_password_submit).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Credentials"

    headers = ["First Name", "Email", "Phone", "Temporary Password"]
    if include_voter_token:
        headers.append("Voter Token")
    headers.append("Account Setup Link")
    sheet.append(headers)

    link_font = Font(color="0F5132", underline="single")
    for row in credentials:
        values = [row["first_name"], row["email"], row["phone"], row["temp_password"]]
        if include_voter_token:
            values.append(row["voter_token"])
        setup_link = row.get("setup_link") or ""
        values.append(setup_link)
        sheet.append(values)

        if setup_link:
            cell = sheet.cell(row=sheet.max_row, column=len(values))
            cell.hyperlink = setup_link
            cell.font = link_font

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class RegisterFileError(Exception):
    """Raised when a file can't be read as a register at all (not just a bad
    row) — e.g. corrupt upload, unreadable encoding, or no recognizable
    header row anywhere near the top of the sheet."""


# Real school/registrar exports rarely use our exact column names. Map every
# header variant we've seen (or are likely to see) onto the field our import
# logic actually needs — matched after normalization (lowercased, non-alnum
# collapsed to single underscores), so "Student ID", "STUDENT_ID", "student
# id" all resolve the same way.
_HEADER_ALIASES: dict[str, str | None] = {
    "student_id": "student_id",
    "studentid": "student_id",
    "student_no": "student_id",
    "student_number": "student_id",
    "id_number": "student_id",
    "index_number": "student_id",
    "matric_number": "student_id",
    "reg_number": "student_id",
    "registration_number": "student_id",
    "s_n": None,  # serial number column — not useful, explicitly ignored
    "email": "email",
    "email_address": "email",
    "institutional_email": "institutional_email",
    "school_email": "institutional_email",
    "personal_email": "personal_email",
    "name": "name",
    "full_name": "name",
    "fullname": "name",
    "student_name": "name",
    "first_name": "first_name",
    "firstname": "first_name",
    "middle_name": "middle_name",
    "middlename": "middle_name",
    "last_name": "last_name",
    "lastname": "last_name",
    "surname": "last_name",
    "phone": "phone",
    "telephone": "phone",
    "phone_number": "phone",
    "mobile": "phone",
    "mobile_number": "phone",
    "program": "program",
    "programme": "program",
    "course": "program",
    "program_category": "program_category",
    "category": "program_category",
    "level": "level",
}

# A header row must contain at least one of these to be recognized as the
# real header row — lets us skip past letterhead/title rows that many school
# exports put above the actual table (institution name, "REGISTRATION
# STATUS", session/semester labels, blank rows, etc.).
_HEADER_SIGNAL_TOKENS = {"student_id", "email", "institutional_email", "personal_email", "name", "first_name"}

# How many leading rows to scan for a real header row before giving up.
_MAX_PREAMBLE_ROWS = 25


def _normalize_header(raw: str) -> str:
    text = re.sub(r"[^a-z0-9]+", "_", str(raw or "").strip().lower()).strip("_")
    return text


def _looks_like_header_row(raw_cells: list) -> bool:
    normalized = {_HEADER_ALIASES.get(_normalize_header(c), _normalize_header(c)) for c in raw_cells}
    return bool(normalized & _HEADER_SIGNAL_TOKENS)


def _merge_row_fields(row: dict) -> dict:
    """Resolve the aliased columns actually present on a row into the plain
    student_id/email/name/phone/program/program_category fields the import
    logic expects — combining first/middle/last name into one name, and
    preferring an institutional email over a personal one when both exist."""
    if not row.get("email"):
        row["email"] = row.get("institutional_email") or row.get("personal_email") or ""

    if not row.get("name"):
        parts = [row.get("first_name"), row.get("middle_name"), row.get("last_name")]
        combined = " ".join(p for p in parts if p)
        if combined:
            row["name"] = combined

    return row


def _rows_from_grid(grid: list[list]) -> list[dict]:
    """Given a 2D grid of raw cell values (already stripped of totally blank
    leading rows), find the real header row within the first few rows, map
    its columns via _HEADER_ALIASES, and build field-normalized row dicts."""
    header_row_index = None
    normalized_headers: list[str | None] = []

    for i, raw_cells in enumerate(grid[:_MAX_PREAMBLE_ROWS]):
        if not any(str(c or "").strip() for c in raw_cells):
            continue  # fully blank row — keep scanning
        if _looks_like_header_row(raw_cells):
            header_row_index = i
            normalized_headers = [
                _HEADER_ALIASES.get(_normalize_header(c), _normalize_header(c)) for c in raw_cells
            ]
            break

    if header_row_index is None:
        raise RegisterFileError(
            "Could not find a header row with a recognizable student ID or email "
            "column anywhere in the first rows of this file. Check the file has a "
            "row naming its columns (e.g. \"Student ID\", \"Email\")."
        )

    rows: list[dict] = []
    for raw_cells in grid[header_row_index + 1 :]:
        if not any(str(c or "").strip() for c in raw_cells):
            continue  # skip blank rows anywhere in the body
        row: dict = {}
        for header, value in zip(normalized_headers, raw_cells):
            if not header:  # unmapped/ignored column (e.g. S/N, gender, hostel...)
                continue
            text = "" if value is None else str(value).strip()
            if text:
                row[header] = text
        if row:
            rows.append(_merge_row_fields(row))

    return rows


def parse_register_file(filename: str, contents: bytes) -> list[dict]:
    """Parse a CSV or .xlsx register upload into student_id/email/name/...
    row dicts, tolerating real-world school export quirks: letterhead rows
    above the real header, differently-named or reordered columns, extra
    columns we don't use, and split first/middle/last name or
    personal/institutional email columns. Raises RegisterFileError if the
    file can't be read at all or has no recognizable header row — never
    raises for an individual bad/incomplete row, which the caller reports
    per-row instead (see import_register / import_members)."""
    suffix = Path(filename or "").suffix.lower()

    try:
        if suffix == ".xlsx":
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            grid = [list(row) for row in sheet.iter_rows(values_only=True)]
        else:
            try:
                text = contents.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = contents.decode("latin-1")
            # Files saved with a .csv extension aren't always comma-delimited —
            # pasting a register out of Excel/Sheets into a text editor commonly
            # produces a tab-separated file that still gets named "register.csv".
            # Sniff the real delimiter instead of assuming comma.
            sample = "\n".join(text.splitlines()[:10])
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            grid = [row for row in csv.reader(io.StringIO(text), dialect) if row]
    except RegisterFileError:
        raise
    except Exception as err:
        raise RegisterFileError(
            f"Could not read this file as a {suffix or 'CSV'} register: {err}"
        ) from err

    return _rows_from_grid(grid)
